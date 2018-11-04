import json
import platform
import re
import shutil
import sys
from io import StringIO
from pathlib import Path

import pexpect
import pexpect.popen_spawn
from openpyxl import load_workbook, Workbook
from tqdm import tqdm

LINE_REGEX = re.compile(r"info depth (?P<depth>\d+) seldepth (?P<seldepth>\d+) time (?P<time>\d+).*nps (?P<nps>\d+)")
MOVE_TIME = 60
PROCESSING_TIME = 63


def check_restarting(config_folder):
    if len(list((config_folder / "processed").glob("*.config"))) == 0:
        return False

    want_to_restart = input(
        "Do you want to restart from where the program left off? y/n/q "
    )
    if want_to_restart in ("q", "Q"):
        sys.exit(0)
    return want_to_restart in ("y", "Y")


def create_config_files(config_folder, options_file):
    prepare_folder(config_folder)
    options = get_config_options(options_file)
    flags = list(options.keys())
    config_options = build_config_options(flags, options, {})
    write_config_files(config_folder, config_options)


def remove_old_results(results_file):
    if results_file.exists():
        results_file.unlink()


def prepare_folder(folder: Path):
    if folder.exists():
        shutil.rmtree(str(folder.absolute()))
    Path.mkdir(folder.absolute(), exist_ok=True)
    Path.mkdir((folder / "processed").absolute(), exist_ok=True)


def get_config_options(config_file):
    config = json.load(open(config_file.absolute()))
    return config["options"]


def get_time_config(config_file):
    config = json.load(open(config_file.absolute()))
    return config["seconds_per_move"]


def build_config_options(flags, options, result):
    if not len(flags):
        return [result]

    calls = []
    flag = flags.pop()
    for option in options[flag]:
        result[flag] = option
        [
            calls.append(i.copy())
            for i in build_config_options(flags.copy(), options, result)
        ]

    return calls


def write_config_files(config_folder, config_options):
    for i, config_values in enumerate(tqdm(config_options, "Generating Configs")):
        with open((config_folder / f"{i+1}.config").absolute(), "w") as f:
            for flag, value in config_values.items():
                f.write(f"--{flag}={value}\n")


def get_results_excel(results_file):
    try:
        wb = load_workbook(str(results_file.absolute()))
    except FileNotFoundError:
        wb = Workbook()

    if 'Results' not in wb.sheetnames:
        ws = wb.create_sheet(title='Results', index=0)
    else:
        ws = wb['Results']

    return wb, ws


def add_results_workbook_headers(worksheet):
    worksheet["A1"] = "LC0 Parameters Run Results"
    worksheet["A2"] = "Filename"
    worksheet["B2"] = "NPS"
    worksheet["C2"] = "DEPTH"
    worksheet["D2"] = "SELDEPTH"

    worksheet["F2"] = "Best NPS"
    worksheet["G2"] = "Filename"
    worksheet["F3"] = "=MAX(B:B)"
    worksheet["G3"] = '=INDIRECT("A" & MATCH(MAX(B:B), B:B, 0))'

    worksheet["F5"] = "Best Depth"
    worksheet["G5"] = "Filename"
    worksheet["F6"] = "=MAX(C:C)"
    worksheet["G6"] = '=INDIRECT("A" & MATCH(MAX(C:C), C:C, 0))'

    worksheet["F8"] = "Best SelDepth"
    worksheet["G8"] = "Filename"
    worksheet["F9"] = "=MAX(D:D)"
    worksheet["G9"] = '=INDIRECT("A" & MATCH(MAX(D:D), D:D, 0))'


def add_results_to_sheet(worksheet, results, filename):
    if results is not None:
        depth, seldepth, nps = (
            results.group('depth'),
            results.group('seldepth'),
            results.group('nps')
        )
    else:
        depth, seldepth, nps = None, None, None

    row = int(filename.split(".")[0])
    worksheet.cell(row + 2, 1, filename)
    worksheet.cell(row + 2, 2, int(nps if nps else 0))
    worksheet.cell(row + 2, 3, int(depth if depth else 0))
    worksheet.cell(row + 2, 4, int(seldepth if seldepth else 0))


def run_lc0_command(lco: Path, config_file: Path, seconds_per_move):
    if platform.system() == "Windows":
        cmd = f'{str(lco.absolute())} -c {str(config_file.relative_to(lco.parent))}'
        child = pexpect.popen_spawn.PopenSpawn(
            cmd,
            timeout=max(PROCESSING_TIME, seconds_per_move*2),
            encoding="utf-8",
        )
    else:
        child = pexpect.spawn(
            str(lco.absolute()),
            ["-c", str(config_file.relative_to(lco.parent))],
            timeout=max(PROCESSING_TIME, seconds_per_move*2),
            encoding="utf-8",
        )

    output = StringIO()
    child.logfile = output
    try:
        child.sendline(f"go infinite movetime {seconds_per_move * 1000}")
        child.expect("bestmove")
        child.sendline("quit")
        child.expect(pexpect.EOF)
    except pexpect.exceptions.TIMEOUT:
        pass

    for line in output.getvalue().splitlines()[::-1]:
        match = LINE_REGEX.match(line)
        if match:
            return match
